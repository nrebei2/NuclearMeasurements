from openpyxl import load_workbook
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

center = [.535,0,-24.2]

def sqauredDist(a,b):
    return ((a[0] - b[0]) ** 2) + ((a[1] - b[1]) ** 2) + ((a[2] - b[2]) ** 2)

Points = []
with open("Points", "r") as data:
    for line in data:
        Points.append([float(i) for i in line.split()])

Order = []
with open("Order", "r") as data:
    for line in data:
        Order.append(int(line)-1)

# Plot Points
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')

for i in range(len(Points)): # plot each point + it's index as text above
    x = Points[i][0]
    y = Points[i][1]
    z = Points[i][2]
    label = Order[i] + 1
    ax.scatter(x, y, z, color='b')
    ax.text(x, y, z, '%s' % (label), size=20, zorder=1, color='k')

ax.set_xlabel('x')
ax.set_ylabel('y')
ax.set_zlabel('z')
plt.show()

# Rotate Plot
# for angle in range(0, 360):
#   ax.view_init(30, angle)
#   plt.draw()
#   plt.pause(.001)

# Normalize Distance
# ratios = []
# for point in Points:
#     ratios.append(sqauredDist(point, center)/sqauredDist(center, Points[27]))
#
# ratios = [ratios[i] for i in Order]
#
# Sheets = ["corner1","corner2"]
#
# import pandas as pd
# for sheet in Sheets:
#     xl = pd.ExcelFile('../../Data/PipeSource.xlsx')
#
#     Counts = []
#     Counts.extend(list(xl.parse(sheet).iloc[:, 0]))
#
#     NewCounts = [a*b for a, b in zip(Counts, ratios)]
#
#     wb = load_workbook('../../Data/PipeSource.xlsx')
#     ws1 = wb[sheet]
#
#     for i in range(len(NewCounts)):
#         ws1.cell(row=i+2, column=1).value = NewCounts[i]
#
#     wb.save('../../Data/PipeSource.xlsx')

