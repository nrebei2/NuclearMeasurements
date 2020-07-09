import matplotlib
import numpy as np
import scipy.optimize as op
import pandas as pd
from itertools import chain
import sys
import matplotlib.pyplot as plt
from matplotlib import ticker
from matplotlib.colors import LogNorm
from mpl_toolkits.axes_grid1 import make_axes_locatable
from sklearn.preprocessing import StandardScaler
from sklearn import preprocessing

from openpyxl import load_workbook,Workbook

np.set_printoptions(threshold=sys.maxsize)

Out = []

Source = input("What source would you like to use? (pointSource, source1, source2, or PipeSource): ")

xl = pd.ExcelFile('../Data/' + Source + '.xlsx')

def fillOut(Out, sheet):
    df = xl.parse(sheet)
    Out.extend(list(np.asarray(df.iloc[:, 0])))
    # remove last element (det 36)
    # del Out[-1]

if Source == "PipeSource":
    #Sheets = ["17.5", "18.5", "19.5", "newpos1", "newpos2", "corner1", "newpos3", "newpos4", "25o5", "25o6", "25o7", "25o8", "newpos5", "newpos6", "corner2", "newpos7", "newpos8", "31.0"]
    #Sheets = ["17.5", "18.5", "19.5", "25o5", "25o6", "25o7", "25o8", "31.0"]
    Sheets = ["25o5", "25o6", "25o7", "25o8"]
    for i in range(len(Sheets)):
        fillOut(Out, Sheets[i])

else:
    Sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 'Sheet6']

    fillOut(Out, Sheets[5])
    fillOut(Out, Sheets[4])
    fillOut(Out, Sheets[3])
    fillOut(Out, Sheets[2])
    fillOut(Out, Sheets[1])
    fillOut(Out, Sheets[0])
    fillOut(Out, Sheets[1])
    fillOut(Out, Sheets[2])
    fillOut(Out, Sheets[3])
    fillOut(Out, Sheets[4])
    fillOut(Out, Sheets[5])




Out = np.asarray(Out)
#Out = np.asarray(list(chain.from_iterable(Out)))


# why does dividing this introduce noise? Fixed, had to start minimization with zero
max = max(Out)
Out = Out/np.max(Out)


R = []
xl = pd.ExcelFile('../Data/NewRespMatr.xlsx')

RSheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 'Sheet6', 'Sheet7', 'Sheet8', 'Sheet9', 'Sheet10']

# Add detector 36
# for sheet in Sheets:
#     df = xl.parse(sheet)
#     x = df.iloc[0, :]
#     y=[]
#     for i in range(11,-1,-1):
#         y.append([x[i] for i in range(12*i, 12*(i+1))])
#     y = list(chain.from_iterable(y))
#
#     wb = load_workbook('../Data/NewRespMatr.xlsx')
#     ws1 = wb[sheet]
#
#     for r in range(0, len(y)):
#         ws1.cell(row=37, column=r+1).value = y[r]
#
#     wb.save('../Data/NewRespMatr.xlsx')

def fillResponse(sheet):
    Rx = []
    df = xl.parse(sheet)
    # range(36) for all detectors
    for x in range(36):
        Rx.append(df.iloc[x, :])
    return np.array(Rx).tolist()


R0 = fillResponse(RSheets[0])
R1 = fillResponse(RSheets[1])
R2 = fillResponse(RSheets[2])
R3 = fillResponse(RSheets[3])
R4 = fillResponse(RSheets[4])
R5 = fillResponse(RSheets[5])
R6 = fillResponse(RSheets[6])
R7 = fillResponse(RSheets[7])
R8 = fillResponse(RSheets[8])
R9 = fillResponse(RSheets[9])

# plt.imshow(R0, cmap='Blues')

plt.show()

# # M * N is the number of pixels
N = 12
no_labels = 5  # how many labels to see on axis x and y.
# M, N is divisible by no_labels.
W = 12  # 10 cm

# Info of the pipe
#innerradius = 10.2 / 2  # in cm
#outerradius = 11.4 / 2  # in cm

innerradius = 4.75
outerradius = 5.25

def deleteOutsiders(R):
    zero_pixels = []
    index = 0
    for y in range(0, N):
        for x in range(0, 12):
            x_p = -5.5 + (x) * 1
            y_p = -5.5 + (y) * 1
            if (x_p) ** 2 + (y_p) ** 2 > innerradius ** 2:
                zero_pixels.append((x, y))
                for k in range(0, 35):
                    del R[k][index]
            else:
                index += 1
    return R


def changeOutsiders(R):
    zero_pixels = []
    G = np.zeros((12, 12))
    for y in range(0, 12):
        y_p = y - 5.5
        for x in range(0, 12):
            x_p = x - 5.5
            if (x_p) ** 2 + (y_p) ** 2 <= innerradius ** 2:
                G[x, y] = 1
            else:
                # G[x,y] = 0
                zero_pixels.append((x, y))
    G = np.ravel(G)
    return np.multiply(R, G)


WantToGraph = input("Do You Want to Graph? (Y or N): ").lower()

if WantToGraph in ["yes", "y"]:
    WantToGraph = True
else:
    WantToGraph = False

if WantToGraph:
    LogScale = input("Do You Want to Color each plot in LogScale? (Y or N): ").lower()

    if LogScale in ["yes", "y"]:
        LogScale = True
    else:
        LogScale = False

Method = input("What method would you like to use? (trf, bvls, fista, or ADAM): ").lower()

if not WantToGraph:
    R0 = np.array(deleteOutsiders(R0))
    R1 = np.array(deleteOutsiders(R1))
    R2 = np.array(deleteOutsiders(R2))
    R3 = np.array(deleteOutsiders(R3))
    R4 = np.array(deleteOutsiders(R4))
    R5 = np.array(deleteOutsiders(R5))
    R6 = np.array(deleteOutsiders(R6))
    R7 = np.array(deleteOutsiders(R7))
    R8 = np.array(deleteOutsiders(R8))
    R9 = np.array(deleteOutsiders(R9))
    Zero = np.zeros((R0.shape[0], 80))

if WantToGraph:
    R0 = np.array(changeOutsiders(R0))
    R1 = np.array(changeOutsiders(R1))
    R2 = np.array(changeOutsiders(R2))
    R3 = np.array(changeOutsiders(R3))
    R4 = np.array(changeOutsiders(R4))
    R5 = np.array(changeOutsiders(R5))
    R6 = np.array(changeOutsiders(R6))
    R7 = np.array(changeOutsiders(R7))
    R8 = np.array(changeOutsiders(R8))
    R9 = np.array(changeOutsiders(R9))
    Zero = np.zeros((R0.shape[0], 144))

# R = np.vstack((np.hstack((R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),
#                np.hstack((Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),
#                np.hstack((Zero, Zero, R9, R8, R7, R6,  R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),
#                np.hstack((Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),
#                np.hstack((Zero, Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero, Zero, Zero, Zero)),
#                np.hstack((Zero, Zero, Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero, Zero, Zero)),
#                 np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero, Zero)),
#                 np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero, Zero)),
#                 np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero, Zero)),
#                 np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero)),
#                 np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9))))

# Assuming we do not know the shape of the pipe after our count measurements
# I think its better since:
# 1. Solution seems to be very similar to above response matrix
# 2. With the above response matrix, we are trying to find counts that are too far out from our original source
    # measurement, resulting in the solution for the above response matrix always assuming most of the holdup is
    # where we measured it, i.e. -5 <= z <= 5 while the activity drops off the further out, meaning the solution is
    # incorrect the further away from 0 where the counts was the greatest. This can be seen in how the solution for
    # source1 drops after z=2. This can be fixed a bit by stopping the minimization from going on too long after solving.

if Source == "PipeSource":
    R = np.vstack((np.hstack((R0, R1, R2, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # 17.5
                   np.hstack((R1, R0, R1, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # 18.5
                   np.hstack((R2, R1, R0, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # 19.5
                   np.hstack((Zero, Zero, Zero, R0, R1, R2, R3, R4, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # newpos1
                   np.hstack((Zero, Zero, Zero, R1, R0, R1, R2, R3, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # newpos2
                   np.hstack((Zero, Zero, Zero, R2, R1, R0, R1, R2, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # corner1
                   np.hstack((Zero, Zero, Zero, R3, R2, R1, R0, R1, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # newpos3
                   np.hstack((Zero, Zero, Zero, R4, R3, R2, R1, R0, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero)),  # newpos4
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R0, R1, R2, R3, Zero, Zero, Zero, Zero, Zero, Zero)),  # 25o5
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R1, R0, R1, R2, Zero, Zero, Zero, Zero, Zero, Zero)),  # 25o6
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R2, R1, R0, R1, Zero, Zero, Zero, Zero, Zero, Zero)),  # 25o7
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R3, R2, R1, R0, Zero, Zero, Zero, Zero, Zero, Zero)),  # 25o8
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R0, R1, R2, R3, R4, Zero)),  # newpos5
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R1, R0, R1, R2, R3, Zero)),  # newpos6
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R2, R1, R0, R1, R2, Zero)),  # corner2
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R3, R2, R1, R0, R1, Zero)),  # newpos7
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R4, R3, R2, R1, R0, Zero)),  # newpos8
                   np.hstack((Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, Zero, R0))))  # 31.0

    R = np.vstack((np.hstack((R0, R1, R2, R3)),
                   np.hstack((R1, R0, R1, R2)),
                   np.hstack((R2, R1, R0, R1)),
                   np.hstack((R3, R2, R1, R0))))

    # R = np.vstack((np.hstack((R0, R1, R2, R3, R4, R5, R6, R7)),
    #              np.hstack((R1, R0, R1, R2, R3, R4, R5, R6)),
    #               np.hstack((R2, R1, R0, R1, R2, R3, R4, R5)),
    #               np.hstack((R3, R2, R1, R0, R1, R2, R3, R4)),
    #               np.hstack((R4, R3, R2, R1, R0, R1, R2, R3)),
    #               np.hstack((R5, R4, R3, R2, R1, R0, R1, R2)),
    #               np.hstack((R6, R5, R4, R3, R2, R1, R0, R1)),
    #               np.hstack((R7, R6, R5, R4, R3, R2, R1, R0))))

    # R = np.vstack((np.hstack((R0, R1, R2, R7, R8, R9, Zero, Zero)),
    #                np.hstack((R1, R0, R1, R6, R7, R8, R9, Zero)),
    #                np.hstack((R2, R1, R0, R5, R6, R7, R8, R9)),
    #                np.hstack((R7, R6, R5, R0, R1, R2, R3, R4)),
    #                np.hstack((R8, R7, R6, R1, R0, R1, R2, R3)),
    #                np.hstack((R9, R8, R7, R2, R1, R0, R1, R2)),
    #                np.hstack((Zero, R9, R8, R3, R2, R1, R0, R1)),
    #                np.hstack((Zero, Zero, R9, R8, R7, R6, R5, R0))))

else:
    R = np.vstack((np.hstack((R0, R1, R2, R3, R4, R5, R6, R7, R8, R9, Zero)),
                    np.hstack((R1, R0, R1, R2, R3, R4, R5, R6, R7, R8, R9)),
                    np.hstack((R2, R1, R0, R1, R2, R3, R4, R5, R6, R7, R8)),
                    np.hstack((R3, R2, R1, R0, R1, R2, R3, R4, R5, R6, R7)),
                    np.hstack((R4, R3, R2, R1, R0, R1, R2, R3, R4, R5, R6)),
                    np.hstack((R5, R4, R3, R2, R1, R0, R1, R2, R3, R4, R5)),
                    np.hstack((R6, R5, R4, R3, R2, R1, R0, R1, R2, R3, R4)),
                    np.hstack((R7, R6, R5, R4, R3, R2, R1, R0, R1, R2, R3)),
                    np.hstack((R8, R7, R6, R5, R4, R3, R2, R1, R0, R1, R2)),
                    np.hstack((R9, R8, R7, R6, R5, R4, R3, R2, R1, R0, R1)),
                    np.hstack((Zero, R9, R8, R7, R6, R5, R4, R3, R2, R1, R0))))

if WantToGraph:
    fig, ax = plt.subplots(6, 6, figsize=(6, 6))
    dets = []
    for i in range(101, 101 + R0.shape[0]):
        dets.append(i)

    k = 0
    for j in range(6):
        for i in range(6):
            matfig = ax[i][j].imshow(R0[k].reshape(12, 12), extent=[-6, 6, -6, 6], origin='lower')
            # # create an axes on the right side of ax. The width of cax will be 5%
            # # of ax and the padding between cax and ax will be fixed at 0.05 inch.
            # divider = make_axes_locatable(ax[i][j])
            # cax = divider.append_axes("right", size="5%", pad=0.05)
            # plt.colorbar(matfig, cax=cax)

            # plot the pipe
            innerradius = innerradius / 12 * 12  # in px
            outerradius = outerradius / 12 * 12  # in px
            center = (0, 0)  # in px
            circle1 = plt.Circle(center, innerradius, color='r', fill=False)
            circle2 = plt.Circle(center, outerradius, color='r', fill=False)
            ax[i][j].add_artist(circle1)
            ax[i][j].add_artist(circle2)

            # labels
            ax[i][j].set_xlabel('x (cm)')
            ax[i][j].set_title('Det {}'.format(dets[k]))
            k += 1
            ax[i][0].set_ylabel('y (cm)')
    #fig.tight_layout()
    # plt.show()
    # plt.savefig("matfig")

# Calculating In with a Bounded-Variable Least-Squares or TRF algorithm
if Method in ["bvls", "trf"]:
    In = op.lsq_linear(R, Out, (0, np.inf),
                       method=Method,
                       max_iter=1000,
                       tol=1e-10,
                       verbose=2)['x']

from scipy.sparse.linalg import svds
if Method == "fista":

    # Why use svd to compute the step step-size of fista
    u, s, vt = svds(R, 1, which='LM')  # svd
    L = s[0] ** 2

    # solving
    # initialize
    N = np.shape(R)[1]
    x = z = np.random.rand(N) * 0

    #L = 10

    # ADAM and FISTA producing same results, all that matters is the stepsize and convergence criteria -> FISTA, since less order complexity
    # Overfitting includes sparsing the solution w/ lambda (similar to l1 approximation?), which in addition under-approximates
    # So, you have to find a balance between sparsing the data too little and too much (much is bad as the pixels which should have equal activity become further apart)
    # Kinda difficult with only 3 sources, also hard to cross-validate given you have you actually observe the graphed solution and not just the calculated activity.

    t = 1
    lam = 10E-8  # tune this parameter to get better result
    max_iter = 2000
    results = []
    err = np.zeros(max_iter)
    for k in np.arange(0, max_iter):
        # print(k)
        xp = z - 1 / L * R.T @ (R @ z - Out)
        xp = np.maximum(xp - lam / L, 0).ravel()
        tp = (1 + np.sqrt(1 + 4 * t ** 2)) / 2
        z = xp + (t - 1) / tp * (xp - x)
        # update
        t = tp
        x = xp
        err[k] = np.linalg.norm(Out - R.dot(x))
        if k % 10 == 0:
            results.append(x)
        # Stops loop when error difference is small -- Prevents overfitting??
        if np.abs(err[k] - err[k-1])/err[k] <= 10 ** -4.5: #or err[k] <= 10E-3 and k >= 1:
            break
    In = results[-1]

    fig, ax1 = plt.subplots(1, 1, figsize=(8, 4.5))
    # ax1.set_title("Reconstruction\nFiltered back projection")
    ax1.plot(np.arange(0, max_iter), err)
    # ax1.set_yscale('log')
    ax1.set_xlabel('Iteration')
    ax1.set_ylabel('$||Ax-y||$')

def noisy_val_grad(theta_hat, data_, label_, deg, lamba):
    gradient = np.zeros_like(theta_hat)
    loss = 0

    for i in range(data_.shape[0]):
        x_ = data_[i, :].reshape(-1, 1)
        y_ = label_[i, 0]
        err = np.matmul(np.transpose(x_), theta_hat) - y_
        if theta_hat[i] == 0:
            grad = 2 * (np.matmul(np.transpose(x_), theta_hat) - y_) * x_ * (
                    np.absolute(np.matmul(np.transpose(x_), theta_hat) - y_) ** (deg - 2))
        else:
            grad = 2 * (np.matmul(np.transpose(x_), theta_hat) - y_) * x_ * (
                    np.absolute(np.matmul(np.transpose(x_), theta_hat) - y_) ** (deg - 2)) \
                   + lamba * np.abs(theta_hat[i])/theta_hat[i]
        #print(grad.shape)
        l = np.abs(err) ** deg
        loss += l / data_.shape[0]
        gradient += np.divide(grad, data_.shape[0])

    return loss, gradient

if Method == 'adam':
    Out = np.asarray([[i] for i in Out])
    lr = 1
    max_iter = 1000
    theta_init = np.random.random((np.shape(R)[1], 1)) * 0
    deg_ = 2
    data_num = np.shape(R)[0]
    batch_size = data_num
    beta_1 = 0.9
    beta_2 = 0.999
    m = 0
    nu = 0
    e = 10E-8
    G = 0
    p_inf = 2/(1-beta_1) - 1
    p = 0
    l = 0

    lam = 10E-9
    u, s, vt = svds(R, 1, which='LM')  # svd
    L = s[0] ** 2
    #lr = 1/L

    # initialize momentum
    mew = 0.9999
    v = 0

    In = theta_init.copy()
    err = np.zeros(max_iter)
    stuff = []
    for t in range(max_iter):
        #print(t)
        idx = np.random.choice(data_num, batch_size)
        train_loss, gradient = noisy_val_grad(In, R, Out, deg_, 0)
        #err[t] = train_loss
        err[t] = np.linalg.norm(Out - R.dot(In))

        # Rectified ADAM
        # m = beta_1 * m + (1 - beta_1) * gradient
        # nu = 1/beta_2 * nu + (1 - beta_2) * np.square(gradient)
        # m_bar = m / (1 - beta_1 ** (t + 1))
        # p = p_inf - 2 * t * (beta_2 ** (t + 1))/(1 - (beta_2 ** (t + 1)))
        # nu_bar = nu / (1 - beta_2 ** (t + 1))
        # if p > 4:
        #     print("true")
        #     #print(((p_inf - 4)*(p_inf-2)*p))
        #     #print(nu)
        #     In = In - lr * (np.sqrt(((p - 4)*(p-2)*p_inf)/((p_inf - 4)*(p_inf-2)*p))) * np.sqrt((1-(beta_2 ** (t + 1)))/(nu+e)) # m_bar / (np.sqrt(nu_bar) + e)
        # else:
        #     In = In - lr * m_bar

        #ADAM
        m = beta_1 * m + (1 - beta_1) * gradient
        nu = beta_2 * nu + (1 - beta_2) * np.square(gradient)
        m_bar = m / (1 - beta_1 ** (t + 1))
        nu_bar = nu / (1 - beta_2 ** (t + 1))
        In = In - lr * m_bar / (np.sqrt(nu_bar) + e)

        # ADAGRAD
        # stuff.append(gradient ** 2)
        # G = sum(stuff)
        # In = In - (gradient * (lr / (np.sqrt(G + e))))

        # GD
        In = In - lr * gradient

        # Nesterov accelerated gradient
        # _, gradient = noisy_val_grad(In - mew * v, R, Out, deg_, 0)
        # v = mew * v + lr * gradient
        # In = In - v

        In = np.asarray([[i] for i in np.maximum(In, 0).ravel()])
        # 10E-4 works best for source1
        if np.abs(err[t] - err[t-1])/err[t] <= 10 ** -4: # or err[t] <= 10E-3 and t >= 1:
            break

    fig, ax1 = plt.subplots(1, 1, figsize=(8, 4.5))
    # ax1.set_title("Reconstruction\nFiltered back projection")
    ax1.plot(np.arange(0, max_iter), err)
    # ax1.set_yscale('log')
    ax1.set_xlabel('Iteration')
    ax1.set_ylabel('$||Ax-y||$')
    #print(np.matmul(R,In) - Out)
    #print(np.matmul(R,In) - Out)

In = In*max
print("The activity using " + Method + " is {} Bq".format(sum(In.ravel())))

# Add calculated activity
# Sheets = ['Point Source', 'Source1', 'Source2']
#
# wb = load_workbook('../Data/Pixel Activity Data.xlsx')
# ws1 = wb[Sheets[0]]
#
# coorinates = []
# for y in range(0, 12):
#     for x in range(0, 12):
#         x_p = -5.5 + (x) * 1
#         y_p = -5.5 + (y) * 1
#         coorinates.append("(" + str(x_p) + " " + str(y_p) + ")")
#
# index = 0
# for r in range(0, 11):
#     for pixel in range(1,145):
#         ws1.cell(row=pixel+1, column=r+2).value = In[index]
#         index += 1
#
# for i in range(len(coorinates)):
#     ws1.cell(row=i+2, column=1).value = coorinates[i]
#
# wb.save('../Data/Pixel Activity Data.xlsx')



if WantToGraph:
    In = np.array_split(In, np.shape(R)[1]/144)
    In = np.asarray(In)

    # plt.imshow(In[4].reshape(12,12))

    fig.suptitle(Source + ', ' + Method, fontsize=16)
    if Source == "PipeSource":
        startingz = currentz = 0
        x = 2
        y = 2
    else:
        startingz = currentz = -5
        x = 3
        y = 4

    fig, ax = plt.subplots(x, y, figsize=(10, 10))
    for i in range(x):
        for j in range(y):
            if i == x-1 and j == y-1 and Source != "PipeSource":
                ax[i][j].axis("off")
                continue
            if LogScale:
                matfig = ax[i][j].imshow(np.log(In[currentz + startingz].reshape(12, 12) + 0.001), extent=[-6, 6, -6, 6],
                                         origin='lower', vmin=0, vmax=np.log(np.amax(In)))
            if not LogScale:
                matfig = ax[i][j].imshow(In[currentz + startingz].reshape(12, 12) + 0.001, extent=[-6, 6, -6, 6],
                                         origin='lower', vmin=0, vmax=np.amax(In))

            # plot the pipe
            center = (0, 0)  # in px
            circle1 = plt.Circle(center, innerradius, color='r', fill=False)
            circle2 = plt.Circle(center, outerradius, color='r', fill=False)
            ax[i][j].add_artist(circle1)
            ax[i][j].add_artist(circle2)

            # labels
            ax[i][j].set_xlabel('x (cm)')
            ax[i][j].set_ylabel('y (cm)')
            if Source == "PipeSource":
                ax[i][j].set_title(Sheets[currentz])
            else:
                ax[i][j].set_title('z = {} cm'.format(currentz))

            fig.tight_layout()
            currentz += 1

    fig.subplots_adjust(right=0.8, top=0.93)
    # put colorbar at desire position
    cbar_ax = fig.add_axes([0.85, 0.15, 0.05, 0.7])
    fig.colorbar(matfig, cax=cbar_ax)
    fig.suptitle(Source + ', ' + Method, fontsize=16)
    plt.show()