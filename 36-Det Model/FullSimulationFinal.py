from math import ceil, sqrt
import matplotlib
import numpy as np
import scipy.optimize as op
import pandas as pd
from itertools import chain
import sys
import matplotlib.pyplot as plt
from matplotlib import ticker
from matplotlib.colors import LogNorm
from mpl_toolkits.axes_grid1 import make_axes_locatable
from sklearn.preprocessing import StandardScaler
from sklearn import preprocessing
from openpyxl import load_workbook,Workbook

# values to change:
numberOfNew = 4
iterations = 100

np.set_printoptions(threshold=sys.maxsize)

Out = []

Source = input("What source would you like to use? (pointSource, source1, source2, or PipeSource): ")

xl = pd.ExcelFile('../Data/' + Source + '.xlsx')

if Source == "PipeSource":
    Sheets = ["17.5", "18.5", "19.5", "corner1", "25o5", "25o6", "25o7", "25o8", "corner2", "31.0"]
    UpdatedSheets = Sheets.copy()
    x = 1
    for i in range(len(Sheets)):
        df = xl.parse(Sheets[i])
        Out.extend(list(np.asarray(df.iloc[:, 0])))
        if Sheets[i] in ("19.5", "corner1", "25o8", "corner2"):
            for j in range(numberOfNew):
                j = j+1
                list1 = np.asarray(df.iloc[:, 0])
                df = xl.parse(Sheets[i+1])
                list2 = np.asarray(df.iloc[:, 0])
                Out.extend(list(list1 + (list2 - list1) * j/(numberOfNew+1)))
                for k in range(len(UpdatedSheets)):
                    if UpdatedSheets[k] == Sheets[i]:
                        UpdatedSheets.insert(k+1, 'newpos' + str((x)*numberOfNew-j+1))
                        break
            x += 1
    Sheets = UpdatedSheets
else:
    Sheets = ['Sheet6', 'Sheet5', 'Sheet4', 'Sheet3', 'Sheet2', 'Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 'Sheet6']

    for i in range(len(Sheets)):
        df = xl.parse(Sheets[i])
        Out.extend(list(np.asarray(df.iloc[:, 0])))
        # to remove last element (det 36)
        #del Out[-1]

Out = np.asarray(Out)

max = max(Out)
Out = Out/np.max(Out)

R = []
xl = pd.ExcelFile('../Data/NewRespMatr.xlsx')

RSheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 'Sheet6', 'Sheet7', 'Sheet8', 'Sheet9', 'Sheet10']

def fillResponse(sheet):
    Rx = []
    df = xl.parse(sheet)
    # range(36) for all detectors
    if Source == 'pointSource':
        numDet = 35
    else:
        numDet = 36
    for x in range(numDet):
        Rx.append(df.iloc[x, :])
    return np.array(Rx).tolist()


R0 = fillResponse(RSheets[0])
R1 = fillResponse(RSheets[1])
R2 = fillResponse(RSheets[2])
R3 = fillResponse(RSheets[3])
R4 = fillResponse(RSheets[4])
R5 = fillResponse(RSheets[5])
R6 = fillResponse(RSheets[6])
R7 = fillResponse(RSheets[7])
R8 = fillResponse(RSheets[8])
R9 = fillResponse(RSheets[9])

# # M * N is the number of pixels
N = 12
no_labels = 5  # how many labels to see on axis x and y.
# M, N is divisible by no_labels.
W = 12  # 10 cm

# Info of the pipe

innerradius = 4.75
outerradius = 5.25

def changeOutsiders(R):
    zero_pixels = []
    G = np.zeros((12, 12))
    for y in range(0, 12):
        y_p = y - 5.5
        for x in range(0, 12):
            x_p = x - 5.5
            if (x_p) ** 2 + (y_p) ** 2 <= innerradius ** 2:
                G[x, y] = 1
            else:
                # G[x,y] = 0
                zero_pixels.append((x, y))
    G = np.ravel(G)
    return np.multiply(R, G)

LogScale = input("Do You Want to Color each plot in LogScale? (Y or N): ").lower()

if LogScale in ["yes", "y"]:
    LogScale = True
else:
    LogScale = False

R0 = np.array(changeOutsiders(R0))
R1 = np.array(changeOutsiders(R1))
R2 = np.array(changeOutsiders(R2))
R3 = np.array(changeOutsiders(R3))
R4 = np.array(changeOutsiders(R4))
R5 = np.array(changeOutsiders(R5))
R6 = np.array(changeOutsiders(R6))
R7 = np.array(changeOutsiders(R7))
R8 = np.array(changeOutsiders(R8))
R9 = np.array(changeOutsiders(R9))
Zero = np.zeros((R0.shape[0], 144))


R = []
length = len(Sheets)
for j in range(length):
    R.append(np.hstack([Zero] * (j-9) + ([R9, R8, R7, R6, R5, R4, R3, R2, R1][-j:] if j > 0 else []) + [R0, R1, R2, R3, R4, R5, R6, R7, R8, R9][0:length-j] +
                       [Zero] * (length-10-j)))
R = np.vstack(R)
#
# fig, ax = plt.subplots(6, 6, figsize=(6, 6))
# dets = []
# for i in range(101, 101 + R0.shape[0]):
#     dets.append(i)
#
# k = 0
# for j in range(6):
#     for i in range(6):
#         matfig = ax[i][j].imshow(R0[k].reshape(12, 12), extent=[-6, 6, -6, 6], origin='lower')
#         # # create an axes on the right side of ax. The width of cax will be 5%
#         # # of ax and the padding between cax and ax will be fixed at 0.05 inch.
#         # divider = make_axes_locatable(ax[i][j])
#         # cax = divider.append_axes("right", size="5%", pad=0.05)
#         # plt.colorbar(matfig, cax=cax)
#
#         # plot the pipe
#         innerradius = innerradius / 12 * 12  # in px
#         outerradius = outerradius / 12 * 12  # in px
#         center = (0, 0)  # in px
#         circle1 = plt.Circle(center, innerradius, color='r', fill=False)
#         circle2 = plt.Circle(center, outerradius, color='r', fill=False)
#         ax[i][j].add_artist(circle1)
#         ax[i][j].add_artist(circle2)
#
#         # labels
#         ax[i][j].set_xlabel('x (cm)')
#         ax[i][j].set_title('Det {}'.format(dets[k]))
#         k += 1
#         ax[i][0].set_ylabel('y (cm)')


from scipy.sparse.linalg import svds
# Why use svd to compute the step step-size of fista
u, s, vt = svds(R, 1, which='LM')  # svd

L = s[0] ** 2

# solving
# initialize
N = np.shape(R)[1]
x = z = np.random.rand(N) * 0

t = 1
lam = 10E-8  # tune this parameter to get better result
max_iter = 2000
results = []
mse = []
err = np.zeros(max_iter)
mse = np.zeros(max_iter)
for k in np.arange(0, max_iter):
    # print(k)
    xp = z - 1 / L * R.T @ (R @ z - Out)
    xp = np.maximum(xp - lam / L, 0).ravel()
    tp = (1 + np.sqrt(1 + 4 * t ** 2)) / 2
    z = xp + (t - 1) / tp * (xp - x)
    # update
    t = tp
    x = xp
    err[k] = np.linalg.norm(Out - R.dot(x))
    #mse[k] = ((BaseIn - x)**2).mean(axis=None)
    if k % 10 == 0:
        results.append(x)
    # Stops loop when error difference is small -- Prevents overfitting??
    if np.abs(err[k] - err[k-1])/err[k] <= 10 ** -4: #or err[k] <= 10E-3 and k >= 1:
        break
    if k == iterations:
        break
In = results[-1]

fig, ax1 = plt.subplots(1, 1, figsize=(8, 4.5))
#ax1.set_title("Reconstruction\nFiltered back projection")
ax1.plot(np.arange(0, max_iter), err)
# ax1.set_yscale('log')
ax1.set_xlabel('Iteration')
ax1.set_ylabel('$||Ax-y||$')

In = In*max
print("The activity using fista is {} Bq".format(sum(In.ravel())))


#Add calculated activity

wb = load_workbook('../Data/Pixel Activity Data.xlsx')
ws1 = wb['Source2']

coorinates = []
for y in range(0, 12):
    for x in range(0, 12):
        x_p = -5.5 + (x) * 1
        y_p = -5.5 + (y) * 1
        coorinates.append("(" + str(x_p) + " " + str(y_p) + ")")

index = 0
for r in range(0, len(Sheets)):
    ws1.cell(row=1, column=r+2).value = Sheets[r]
    for pixel in range(1,145):
        ws1.cell(row=pixel+1, column=r+2).value = In[index]
        index += 1

for i in range(len(coorinates)):
    ws1.cell(row=i+2, column=1).value = coorinates[i]

wb.save('../Data/Pixel Activity Data.xlsx')




In = np.array_split(In, np.shape(R)[1]/144)
In = np.asarray(In)

fig.suptitle(Source + ', fista', fontsize=16)
startingz = currentz = 0
x = ceil(sqrt(len(Sheets)))
y = ceil(sqrt(len(Sheets)))
# else:
#     startingz = 5
#     currentz = -5
#     x = 3
#     y = 4

fig, ax = plt.subplots(x, y, figsize=(10, 10))
for i in range(x):
    for j in range(y):
        if currentz >= len(Sheets):
            ax[i][j].axis("off")
            continue

        if LogScale:
            matfig = ax[i][j].imshow(np.log(In[currentz + startingz].reshape(12, 12) + 0.00001), extent=[-6, 6, -6, 6],
                                     origin='lower', vmin=0, vmax=np.log(np.amax(In)))
        if not LogScale:
            matfig = ax[i][j].imshow(In[currentz + startingz].reshape(12, 12), extent=[-6, 6, -6, 6],
                                     origin='lower', vmin=0, vmax=np.amax(In))

        # plot the pipe
        center = (0, 0)  # in px
        circle1 = plt.Circle(center, innerradius, color='r', fill=False)
        circle2 = plt.Circle(center, outerradius, color='r', fill=False)
        ax[i][j].add_artist(circle1)
        ax[i][j].add_artist(circle2)
        ax[i][j].axis('off')

        # labels
        # ax[i][j].set_xlabel('x (cm)')
        # ax[i][j].set_ylabel('y (cm)')

        ax[i][j].set_title(Sheets[currentz])

        #fig.tight_layout()
        currentz += 1

fig.subplots_adjust(right=0.8, top=0.93)
# put colorbar at desire position
cbar_ax = fig.add_axes([0.85, 0.15, 0.05, 0.7])
fig.colorbar(matfig, cax=cbar_ax)
fig.suptitle(Source + ', fista', fontsize=16)
plt.show()